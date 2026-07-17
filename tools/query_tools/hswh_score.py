# -*- coding: utf-8 -*-
import json
import os
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webdriver import WebDriver

from tools.registry import register_tool

SITE_URL = "https://hswhds.com/"
API_BASE = "https://api.hswhds.com/api"

OUTPUT_COLUMNS = [
    "ID", "账号", "查询状态", "错误信息", "姓名", "证件号",
    "考生编号", "赛道", "组别", "报名时间", "市赛结果", "发证时间", "省赛", "国赛",
]

# Terminal colors
_GREEN = "\033[92m"
_RED = "\033[91m"
_YELLOW = "\033[93m"
_RESET = "\033[0m"


@register_tool("query_tools", "成绩查询", "hswh_score", "红色文化大赛成绩查询",
               "查询比赛成绩，支持单个、按编号和批量查询", icon="search")
class HswhScoreQuery:
    """红色文化大赛成绩查询工具"""

    def __init__(self, driver_path: str = "", chrome_binary: str = ""):
        self.driver_path = driver_path
        self.chrome_binary = chrome_binary

    # ---- Chrome / ChromeDriver resolution ----

    def _get_chrome_binary(self) -> str:
        if self.chrome_binary:
            return self.chrome_binary
        env_bin = os.environ.get("CHROME_BIN", "")
        if env_bin and Path(env_bin).exists():
            return env_bin
        for p in [
            "/usr/bin/google-chrome",
            "/usr/bin/google-chrome-stable",
            "/usr/bin/chromium",
            "/usr/bin/chromium-browser",
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        ]:
            if Path(p).exists():
                return p
        return ""

    def _get_driver_path(self) -> str:
        if self.driver_path and Path(self.driver_path).exists():
            return self.driver_path
        env_path = os.environ.get("CHROMEDRIVER_PATH", "")
        if env_path and Path(env_path).exists():
            return env_path
        for p in [
            "/usr/bin/chromedriver",
            "/usr/local/bin/chromedriver",
            "/app/chromedriver/chromedriver",
            str(Path(__file__).parent.parent.parent / "chromedriver" / "chromedriver"),
        ]:
            if Path(p).exists():
                return p
        try:
            from tools.query_tools.chromedriver_manager import ensure_chromedriver
            path = ensure_chromedriver()
            if path and Path(path).exists():
                return path
        except Exception:
            pass
        import shutil
        w = shutil.which("chromedriver")
        if w:
            return w
        raise RuntimeError("未找到 ChromeDriver")

    def _make_driver(self) -> WebDriver:
        options = Options()
        options.page_load_strategy = "eager"
        binary = self._get_chrome_binary()
        if binary:
            options.binary_location = binary
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1440,1100")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--lang=zh-CN")
        service = Service(self._get_driver_path())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(20)
        return driver

    # ---- Selenium helpers ----

    def _fetch_json(self, driver: WebDriver, path: str, payload: Optional[Dict], token: str = "") -> Dict:
        url = f"{API_BASE}{path}"
        script = """
            const [url, payload, token, done] = arguments;
            const headers = {"Content-Type": "application/json", "XX-Device-Type": "wxapp"};
            if (token) headers["XX-Token"] = token;
            fetch(url, {method:"POST", headers, body: JSON.stringify(payload||{})})
            .then(async r => { try { done(JSON.parse(await r.text())); } catch(e) { done({code:-1,msg:"Invalid JSON"}); } })
            .catch(e => done({code:-1, msg:String(e)}));
        """
        result = driver.execute_async_script(script, url, payload or {}, token)
        return result if isinstance(result, dict) else {"code": -1, "msg": str(result)}

    def _login(self, driver: WebDriver, mobile: str, password: str) -> Dict:
        driver.get(SITE_URL)
        driver.execute_script("localStorage.clear(); sessionStorage.clear();")
        resp = self._fetch_json(driver, "/user/login/login", {"mobile": mobile, "password": password, "isagree": 1})
        if resp.get("code") != 1:
            raise RuntimeError(resp.get("msg") or f"登录失败: {resp}")
        data = resp.get("data") or {}
        token = data.get("token") or ""
        if not token:
            raise RuntimeError(f"登录响应没有 token: {resp}")
        driver.execute_script(
            "localStorage.setItem('token',arguments[0]);localStorage.setItem('userInfo',JSON.stringify(arguments[1]||{}));",
            token, data.get("user_info") or {},
        )
        return {"token": token, "user_info": data.get("user_info") or {}}

    def _query_keyword(self, driver: WebDriver, keyword: str) -> List[Dict]:
        resp = self._fetch_json(driver, "/home/index/query", {"keywords": keyword})
        if resp.get("code") != 1:
            raise RuntimeError(resp.get("msg") or f"查询失败: {resp}")
        data = resp.get("data") or []
        return data if isinstance(data, list) else [data]

    def _query_account(self, driver: WebDriver, mobile: str, password: str) -> Dict:
        auth = self._login(driver, mobile, password)
        signup_resp = self._fetch_json(driver, "/user/signup/index", {}, auth["token"])
        if signup_resp.get("code") != 1:
            raise RuntimeError(signup_resp.get("msg") or f"读取报名信息失败")
        signups = (signup_resp.get("data") or {}).get("list") or []
        results = []
        for s in signups:
            no = str(s.get("signup_no") or "").strip()
            results.append({"signup": s, "score_results": self._query_keyword(driver, no) if no else []})
        return {"user_info": auth.get("user_info") or {}, "signups": signups, "results": results}

    # ---- Result flattening ----

    def _pick(self, r: Dict, *keys, default=""):
        for k in keys:
            v = r.get(k)
            if v not in (None, ""):
                return str(v)
        return default

    def _pick_result(self, r: Dict, *keys, default=""):
        v = self._pick(r, *keys, default="")
        return default if v in ("", "0", "None") else v

    def _flatten(self, data, signup_ctx=None) -> List[Dict[str, str]]:
        rows = []
        iterable = data if isinstance(data, list) else data.get("results", []) if isinstance(data, dict) else []
        for item in iterable:
            if not isinstance(item, dict):
                continue
            if "score_results" in item:
                signup = item.get("signup") or {}
                scores = item.get("score_results") or []
                if not scores:
                    rows.append({
                        "姓名": "", "证件号": "",
                        "考生编号": self._pick(signup, "signup_no"),
                        "赛道": self._pick(signup, "road_name"),
                        "组别": self._pick(signup, "groups_name"),
                        "报名时间": self._pick(signup, "time"),
                        "市赛结果": self._pick_result(signup, "has_prize_level_1", default="待公示"),
                        "发证时间": "",
                        "省赛": self._pick_result(signup, "has_prize_level_2") or ("已开放提交" if self._pick(signup, "p_url") else "未开始"),
                        "国赛": self._pick_result(signup, "has_prize_level_3", default="未开始"),
                    })
                for s in scores:
                    rows.extend(self._flatten([s], signup_ctx=signup))
            else:
                rows.append({
                    "姓名": self._pick(item, "user_login", "name"),
                    "证件号": self._pick(item, "card_no"),
                    "考生编号": self._pick(item, "signup_no"),
                    "赛道": self._pick(item, "road_name"),
                    "组别": self._pick(item, "groups_name"),
                    "报名时间": self._pick(item, "time"),
                    "市赛结果": self._pick_result(item, "prize_level_1", "has_prize_level_1", default="待公示"),
                    "发证时间": self._pick(item, "prize_level_1_date"),
                    "省赛": self._pick_result(item, "prize_level_2", "has_prize_level_2") or ("已开放提交" if self._pick(item or {}, "p_url") or self._pick(signup_ctx or {}, "p_url") else "未开始"),
                    "国赛": self._pick_result(item, "prize_level_3", "has_prize_level_3", default="未开始"),
                })
        return rows

    # ---- Single / keyword queries ----

    def query_single(self, mobile: str, password: str) -> Dict:
        driver = self._make_driver()
        try:
            data = self._query_account(driver, mobile, password)
            return {"success": True, "user_info": data.get("user_info"), "results": self._flatten(data)}
        except Exception as e:
            return {"success": False, "error": str(e)}
        finally:
            driver.quit()

    def query_by_keyword(self, keyword: str) -> Dict:
        driver = self._make_driver()
        try:
            driver.get(SITE_URL)
            return {"success": True, "results": self._flatten(self._query_keyword(driver, keyword))}
        except Exception as e:
            return {"success": False, "error": str(e)}
        finally:
            driver.quit()

    # ---- Multi-threaded batch query ----

    def _query_one_account(self, account: Dict, index: int, total: int) -> List[Dict]:
        """Query a single account in its own driver. Thread-safe."""
        mobile = account.get("account", "")
        password = account.get("password", "")
        account_id = account.get("id", str(index))
        tag = f"[{index}/{total}]"

        if not mobile or not password:
            print(f"{_YELLOW}{tag} 跳过: 账号或密码为空 ({mobile}){_RESET}", flush=True)
            return [{"ID": account_id, "账号": mobile, "查询状态": "跳过", "错误信息": "账号或密码为空"}]

        print(f"{_YELLOW}{tag} 查询中: {mobile}{_RESET}", flush=True)
        driver = self._make_driver()
        try:
            data = self._query_account(driver, mobile, password)
            rows = self._flatten(data)
            result = []
            for row in rows:
                result.append({"ID": account_id, "账号": mobile, "查询状态": "成功", "错误信息": "", **row})
            print(f"{_GREEN}{tag} 成功: {mobile} -> {len(rows)} 条记录{_RESET}", flush=True)
            return result
        except Exception as e:
            print(f"{_RED}{tag} 失败: {mobile} -> {e}{_RESET}", flush=True)
            return [{"ID": account_id, "账号": mobile, "查询状态": "失败", "错误信息": str(e)}]
        finally:
            driver.quit()

    def query_batch(self, accounts: List[Dict], progress_callback: Optional[Callable] = None, max_workers: int = 10) -> Dict:
        """
        Multi-threaded batch query.

        Args:
            accounts: list of {id, account, password}
            progress_callback: called as progress_callback(completed, total, partial_results)
            max_workers: number of threads (default 10)
        """
        total = len(accounts)
        all_results = []
        completed = 0
        lock = threading.Lock()

        print(f"\n{'='*50}", flush=True)
        print(f"开始批量查询: 共 {total} 个账号, {max_workers} 个线程", flush=True)
        print(f"{'='*50}\n", flush=True)

        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            future_map = {
                pool.submit(self._query_one_account, acct, i + 1, total): i
                for i, acct in enumerate(accounts)
            }
            for future in as_completed(future_map):
                try:
                    rows = future.result()
                except Exception as e:
                    idx = future_map[future]
                    rows = [{"ID": accounts[idx].get("id", ""), "账号": accounts[idx].get("account", ""),
                             "查询状态": "失败", "错误信息": str(e)}]
                with lock:
                    all_results.extend(rows)
                    completed += 1
                    if progress_callback:
                        progress_callback(completed, total, list(all_results))

        print(f"\n{'='*50}", flush=True)
        ok = sum(1 for r in all_results if r.get("查询状态") == "成功")
        fail = sum(1 for r in all_results if r.get("查询状态") == "失败")
        skip = sum(1 for r in all_results if r.get("查询状态") == "跳过")
        print(f"查询完成: 成功 {ok}, 失败 {fail}, 跳过 {skip}, 共 {len(all_results)} 条记录", flush=True)
        print(f"{'='*50}\n", flush=True)

        return {"success": True, "total": total, "results": all_results}


# ---- Excel helpers ----

def read_batch_accounts(xlsx_path: str) -> List[Dict[str, str]]:
    from openpyxl import load_workbook
    COL_MAP = {"ID": "id", "账号": "account", "密码": "password"}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(v).strip() if v else "" for v in next(rows)]
    except StopIteration:
        raise RuntimeError("Excel 表格为空")
    col_idx = {}
    for i, h in enumerate(headers):
        if h in COL_MAP:
            col_idx[COL_MAP[h]] = i
    missing = [n for n, f in COL_MAP.items() if f not in col_idx]
    if missing:
        raise RuntimeError(f"Excel 表头缺少: {', '.join(missing)}，需要 ID、账号、密码")
    accounts = []
    for row in rows:
        vals = list(row)
        item = {
            "id": str(vals[col_idx["id"]]).strip() if col_idx["id"] < len(vals) and vals[col_idx["id"]] else "",
            "account": str(vals[col_idx["account"]]).strip() if col_idx["account"] < len(vals) and vals[col_idx["account"]] else "",
            "password": str(vals[col_idx["password"]]).strip() if col_idx["password"] < len(vals) and vals[col_idx["password"]] else "",
        }
        if any(item.values()):
            accounts.append(item)
    if not accounts:
        raise RuntimeError("Excel 中没有数据行")
    return accounts


def write_template_xlsx(output_path: str) -> str:
    from openpyxl import Workbook
    p = Path(output_path).expanduser().resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "批量查询模板"
    ws.append(["ID", "账号", "密码"])
    ws.append(["示例1", "请填写手机号", "请填写密码"])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    wb.save(p)
    return str(p)
