# -*- coding: utf-8 -*-
import json
import os
import sys
import time
import tempfile
from flask import Flask, Blueprint, render_template, request, jsonify, send_file, Response
from config import Config


def create_app():
    app = Flask(__name__)
    app.config.from_object(Config)
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    os.makedirs(app.config["OUTPUT_FOLDER"], exist_ok=True)

    from tools.registry import get_all_tools

    @app.route("/health")
    def health():
        return {"status": "ok"}

    @app.context_processor
    def inject_tools():
        return dict(all_tools=get_all_tools())

    app.register_blueprint(bp)
    return app


bp = Blueprint("main", __name__)


# ---- Pages ----

@bp.route("/")
def index():
    from tools.registry import get_all_tools
    return render_template("index.html", tools=get_all_tools())


@bp.route("/tool/<tool_name>")
def tool_page(tool_name):
    from tools.registry import get_tool_by_name
    tool = get_tool_by_name(tool_name)
    if tool is None:
        return render_template("404.html"), 404
    return render_template(f"tools/{tool_name}.html", tool=tool)


# ---- API: Single query ----

@bp.route("/api/query/hswh/single", methods=["POST"])
def api_hswh_single():
    data = request.json
    if not data:
        return jsonify({"success": False, "error": "请提供查询数据"}), 400
    mobile = data.get("mobile", "").strip()
    password = data.get("password", "").strip()
    if not mobile or not password:
        return jsonify({"success": False, "error": "请输入账号和密码"}), 400
    try:
        from tools.query_tools.hswh_score import HswhScoreQuery
        q = HswhScoreQuery(driver_path=app.config.get("CHROMEDRIVER_PATH", ""))
        return jsonify(q.query_single(mobile, password))
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ---- API: Keyword query ----

@bp.route("/api/query/hswh/keyword", methods=["POST"])
def api_hswh_keyword():
    data = request.json
    if not data:
        return jsonify({"success": False, "error": "请提供查询数据"}), 400
    keyword = data.get("keyword", "").strip()
    if not keyword:
        return jsonify({"success": False, "error": "请输入编号"}), 400
    try:
        from tools.query_tools.hswh_score import HswhScoreQuery
        q = HswhScoreQuery(driver_path=app.config.get("CHROMEDRIVER_PATH", ""))
        return jsonify(q.query_by_keyword(keyword))
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ---- API: Batch query (SSE streaming) ----

@bp.route("/api/query/hswh/batch", methods=["POST"])
def api_hswh_batch():
    if "file" not in request.files:
        return jsonify({"success": False, "error": "请上传 Excel 文件"}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "error": "请上传 Excel 文件"}), 400

    try:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        file.save(tmp.name)
        tmp_path = tmp.name

        from tools.query_tools.hswh_score import HswhScoreQuery, read_batch_accounts
        accounts = read_batch_accounts(tmp_path)
        os.unlink(tmp_path)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 400

    q = HswhScoreQuery(driver_path=app.config.get("CHROMEDRIVER_PATH", ""))

    def generate():
        import queue
        q_events = queue.Queue()

        def on_progress(completed, total, partial):
            q_events.put({"type": "progress", "completed": completed, "total": total, "results": partial})

        import threading
        def run_query():
            try:
                result = q.query_batch(accounts, progress_callback=on_progress, max_workers=10)
                q_events.put({"type": "done", "data": result})
            except Exception as e:
                q_events.put({"type": "error", "error": str(e)})

        t = threading.Thread(target=run_query, daemon=True)
        t.start()

        while True:
            try:
                evt = q_events.get(timeout=120)
            except queue.Empty:
                yield f"data: {json.dumps({'type': 'error', 'error': '查询超时'})}\n\n"
                break
            yield f"data: {json.dumps(evt, ensure_ascii=False)}\n\n"
            if evt["type"] in ("done", "error"):
                break

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---- API: Template download ----

@bp.route("/api/query/hswh/template")
def api_hswh_template():
    from tools.query_tools.hswh_score import write_template_xlsx
    tmp = tempfile.mktemp(suffix=".xlsx")
    write_template_xlsx(tmp)
    return send_file(tmp, as_attachment=True, download_name="红色文化大赛成绩查询模板.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---- API: Export results as XLSX ----

@bp.route("/api/query/hswh/export", methods=["POST"])
def api_hswh_export():
    data = request.json
    if not data or not data.get("results"):
        return jsonify({"success": False, "error": "没有可导出的数据"}), 400

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    results = data["results"]
    columns = ["ID", "账号", "查询状态", "错误信息", "姓名", "证件号",
               "考生编号", "赛道", "组别", "报名时间", "市赛结果", "发证时间",
               "省赛", "省赛提交链接", "国赛"]

    wb = Workbook()
    ws = wb.active
    ws.title = "成绩查询结果"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    green_font = Font(color="10B981", bold=True)
    red_font = Font(color="EF4444", bold=True)
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            cell.border = border
            if col_name == "查询状态":
                cell.font = green_font if cell.value == "成功" else red_font

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    import tempfile
    tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    date_str = __import__("datetime").date.today().isoformat()
    return send_file(tmp, as_attachment=True, download_name=f"成绩查询结果_{date_str}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---- API: Tools list ----

@bp.route("/api/tools")
def api_tools():
    from tools.registry import get_all_tools
    return jsonify({"code": 200, "data": get_all_tools()})


if __name__ == "__main__":
    app = create_app()
    app.run(debug=True, host="0.0.0.0", port=5001)
